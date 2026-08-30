from __future__ import annotations

import json
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
SNAPSHOT_DIR = DATA_DIR / "snapshots"
INACTIVE_DAYS = 7
MAX_RECENT_DAYS = 30
BROKERS = {
    "puprime": "PU Prime",
    "vantage": "Vantage",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
COL_ALIASES = {
    "date": "opened",
    "user id": "user_id",
    "userid": "user_id",
    "account": "account",
    "account number": "account",
    "name": "name",
    "account owner": "owner",
    "account type": "account_type",
    "platform": "platform",
    "base currency": "currency",
    "profit": "profit",
    "balance": "balance",
    "account equity": "equity",
    "equity": "equity",
    "credit": "credit",
    "account journey": "journey",
    "account journal": "journey",
}


def _norm_header(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _clean_id(value) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.endswith(".0") and text.replace(".", "", 1).replace("-", "", 1).isdigit():
        return text[:-2]
    return text


def _num(value) -> float:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    if not text or text == "-":
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _text(value) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return str(value).strip()


def map_columns(df: pd.DataFrame) -> pd.DataFrame:
    mapped = {}
    for col in df.columns:
        key = _norm_header(col)
        if key in COL_ALIASES:
            mapped[col] = COL_ALIASES[key]
            continue
        if "journey" in key or key.startswith("account jou"):
            mapped[col] = "journey"
    return df.rename(columns=mapped)


def load_excel(path: Path) -> list[dict]:
    df = pd.read_excel(path, sheet_name=0, dtype=object)
    df = map_columns(df)
    needed = {"account", "user_id"}
    missing = needed - set(df.columns)
    if missing:
        raise ValueError(f"Missing columns: {', '.join(sorted(missing))}")

    accounts = []
    for _, row in df.iterrows():
        account = _clean_id(row.get("account"))
        if not account:
            continue
        accounts.append(
            {
                "account": account,
                "user_id": _clean_id(row.get("user_id")),
                "name": _text(row.get("name")),
                "owner": _text(row.get("owner")),
                "account_type": _text(row.get("account_type")),
                "platform": _text(row.get("platform")),
                "currency": _text(row.get("currency")),
                "opened": _text(row.get("opened")),
                "profit": _num(row.get("profit")),
                "balance": _num(row.get("balance")),
                "equity": _num(row.get("equity")),
                "credit": _num(row.get("credit")),
                "journey": _text(row.get("journey")),
            }
        )
    return accounts


def journey_days(journey: str) -> tuple[str, int | None]:
    text = (journey or "").strip().lower()
    if not text or text == "-":
        return "unknown", None
    match = re.search(r"trading in(?:\s+last)?\s+(\d+)", text)
    if match:
        return "trading_in", int(match.group(1))
    match = re.search(r"traded\s+(\d+)\s+days?\s+ago", text)
    if match:
        return "traded_ago", int(match.group(1))
    return "other", None


def is_inactive_recent(journey: str, inactive_days: int = INACTIVE_DAYS, max_days: int = MAX_RECENT_DAYS) -> bool:
    kind, days = journey_days(journey)
    if days is None:
        return False
    if kind == "trading_in":
        return inactive_days < days <= max_days
    if kind == "traded_ago":
        return inactive_days < days < max_days
    return False


def traded_within_month(journey: str, max_days: int = MAX_RECENT_DAYS) -> bool:
    kind, days = journey_days(journey)
    if days is None:
        return False
    return days <= max_days


def no_funds(acc: dict) -> bool:
    return acc["balance"] <= 0


def unique_users(rows: list[dict]) -> list[dict]:
    seen: dict[str, dict] = {}
    for acc in rows:
        uid = acc.get("user_id") or acc["account"]
        if uid not in seen:
            seen[uid] = {
                "user_id": acc.get("user_id") or "-",
                "name": acc.get("name") or "-",
                "accounts_count": 1,
                "accounts": acc["account"],
            }
        else:
            seen[uid]["accounts_count"] += 1
            seen[uid]["accounts"] += f", {acc['account']}"
    return list(seen.values())


def index_by_account(rows: list[dict]) -> dict[str, dict]:
    return {row["account"]: row for row in rows}


def compare(
    today_rows: list[dict],
    yesterday_rows: list[dict] | None,
    inactive_days: int = INACTIVE_DAYS,
) -> dict:
    today = index_by_account(today_rows)
    yesterday = index_by_account(yesterday_rows or [])

    inactive = [a for a in today_rows if is_inactive_recent(a["journey"], inactive_days)]
    unfunded = [
        a for a in today_rows if no_funds(a) and traded_within_month(a["journey"])
    ]
    if yesterday:
        removed = [a for acc, a in yesterday.items() if acc not in today]
        removed_users = unique_users(removed)
        new_accounts = [a for acc, a in today.items() if acc not in yesterday]
        newly_unfunded = [
            a
            for acc, a in today.items()
            if no_funds(a)
            and traded_within_month(a["journey"])
            and acc in yesterday
            and not no_funds(yesterday[acc])
        ]
        journey_changed = [
            {
                **a,
                "old_journey": yesterday[acc]["journey"],
            }
            for acc, a in today.items()
            if acc in yesterday and a["journey"] != yesterday[acc]["journey"]
        ]
    else:
        removed = []
        removed_users = []
        new_accounts = []
        newly_unfunded = []
        journey_changed = []

    by_user: dict[str, list[dict]] = defaultdict(list)
    for acc in today_rows:
        if acc["user_id"]:
            by_user[acc["user_id"]].append(acc)
    duplicates = {uid: accs for uid, accs in by_user.items() if len(accs) > 1}

    return {
        "generated": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "inactive_days": inactive_days,
        "today_count": len(today_rows),
        "yesterday_count": len(yesterday_rows or []),
        "has_previous": bool(yesterday_rows),
        "inactive": inactive,
        "no_funds": unfunded,
        "removed": removed,
        "removed_users": removed_users,
        "new_accounts": new_accounts,
        "newly_unfunded": newly_unfunded,
        "journey_changed": journey_changed,
        "duplicates": duplicates,
    }


def _row_line(acc: dict) -> str:
    name = acc.get("name") or "-"
    return f"{acc['account']} | {name} | UID {acc.get('user_id') or '-'}"


def format_telegram(result: dict, filename: str) -> str:
    unlinked = result.get("removed_users") or []
    lines = [
        "IB Tracker Report",
        f"Broker: {result.get('broker') or '-'}",
        f"File: {filename}",
        f"Time: {result['generated']}",
        f"Accounts today: {result['today_count']}"
        + (f" | previous: {result['yesterday_count']}" if result["has_previous"] else " | first file"),
        "",
        f"1) Inactive ~7 days (not over 1 month): {len(result['inactive'])}",
        f"2) No funds (traded last 30 days): {len(result['no_funds'])}",
        f"3) Unlinked User IDs: {len(unlinked)}",
    ]

    if not result["has_previous"]:
        lines.append("\nFirst snapshot saved. Send the next Excel for day-to-day compare.")

    def block(title: str, rows: list[dict], extra=None, limit: int = 20):
        lines.append(f"\n{title}")
        if not rows:
            lines.append("None")
            return
        for acc in rows[:limit]:
            extra_txt = extra(acc) if extra else ""
            lines.append(f"• {_row_line(acc)}{extra_txt}")
        if len(rows) > limit:
            lines.append(f"... +{len(rows) - limit} more in Excel report")

    block(
        "Inactive ~7 days",
        result["inactive"],
        lambda a: f" | {a['journey'] or '-'}",
    )
    block(
        "No funds (recent traders)",
        result["no_funds"],
        lambda a: f" | bal {a['balance']:.2f} {a.get('currency') or ''}".rstrip(),
    )
    lines.append("\nUnlinked / removed")
    if not unlinked:
        lines.append("None")
    else:
        for row in unlinked[:20]:
            extra = f" | {row['accounts_count']} accounts" if row["accounts_count"] > 1 else ""
            lines.append(f"• UID {row['user_id']} | {row['name']}{extra}")
        if len(unlinked) > 20:
            lines.append(f"... +{len(unlinked) - 20} more in Excel report")
    return "\n".join(lines)


def _write_sheet(writer: pd.ExcelWriter, name: str, rows: list[dict], columns: list[str]) -> None:
    df = pd.DataFrame(rows)
    if df.empty:
        df = pd.DataFrame(columns=columns)
    else:
        keep = [c for c in columns if c in df.columns]
        df = df[keep]
    df.to_excel(writer, sheet_name=name[:31], index=False)
    ws = writer.sheets[name[:31]]
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for idx, col in enumerate(df.columns, 1):
        width = min(max(12, len(str(col)) + 2, *(len(str(v)) + 2 for v in df[col].head(40))), 36)
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_report_xlsx(result: dict, out_path: Path) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    base_cols = [
        "account",
        "user_id",
        "name",
        "owner",
        "platform",
        "account_type",
        "currency",
        "balance",
        "equity",
        "credit",
        "journey",
        "opened",
    ]
    dup_rows = []
    for uid, accs in result["duplicates"].items():
        for acc in accs:
            dup_rows.append({**acc, "same_user_id": uid, "accounts_count": len(accs)})

    summary = pd.DataFrame(
        [
            {"item": "Broker", "value": result.get("broker") or "-"},
            {"item": "Generated", "value": result["generated"]},
            {"item": "Accounts today", "value": result["today_count"]},
            {"item": "Previous file accounts", "value": result["yesterday_count"]},
            {"item": "Inactive ~7 days (not over 1 month)", "value": len(result["inactive"])},
            {"item": "No funds (traded last 30 days)", "value": len(result["no_funds"])},
            {"item": "Unlinked User IDs", "value": len(result.get("removed_users") or [])},
            {"item": "Duplicate User IDs", "value": len(result["duplicates"])},
            {"item": "New accounts", "value": len(result["new_accounts"])},
            {"item": "Newly no funds", "value": len(result["newly_unfunded"])},
            {"item": "Journey changed", "value": len(result["journey_changed"])},
        ]
    )

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        ws = writer.sheets["Summary"]
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 22
        _write_sheet(writer, "Inactive", result["inactive"], base_cols)
        _write_sheet(writer, "No Funds", result["no_funds"], base_cols)
        _write_sheet(
            writer,
            "Unlinked Users",
            result.get("removed_users") or [],
            ["user_id", "name", "accounts_count", "accounts"],
        )
        _write_sheet(writer, "Duplicates", dup_rows, ["same_user_id", "accounts_count", *base_cols])
        _write_sheet(writer, "New Accounts", result["new_accounts"], base_cols)
        _write_sheet(
            writer,
            "Newly No Funds",
            result["newly_unfunded"],
            base_cols,
        )
        _write_sheet(
            writer,
            "Journey Changed",
            result["journey_changed"],
            ["old_journey", *base_cols],
        )
    return out_path


def detect_broker(filename: str, caption: str = "") -> str | None:
    text = f"{filename} {caption}".lower().replace("_", " ").replace("-", " ")
    if "vantage" in text:
        return "vantage"
    if "pu prime" in text or "puprime" in text:
        return "puprime"
    return None


def snapshot_path(broker: str) -> Path:
    return SNAPSHOT_DIR / f"{broker}.json"


def load_snapshot(broker: str) -> list[dict] | None:
    path = snapshot_path(broker)
    if not path.exists():
        return None
    return json.loads(path.read_text(encoding="utf-8"))


def save_snapshot(rows: list[dict], broker: str) -> None:
    path = snapshot_path(broker)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    hist = DATA_DIR / "history" / broker
    hist.mkdir(parents=True, exist_ok=True)
    (hist / f"{date.today().isoformat()}.json").write_text(
        json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def process_file(
    path: Path,
    broker: str,
    inactive_days: int = INACTIVE_DAYS,
) -> tuple[dict, Path]:
    today_rows = load_excel(path)
    if not today_rows:
        raise ValueError("No accounts found in this file.")
    previous = load_snapshot(broker)
    result = compare(today_rows, previous, inactive_days=inactive_days)
    result["broker"] = BROKERS.get(broker, broker)
    result["broker_key"] = broker
    save_snapshot(today_rows, broker)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = DATA_DIR / "reports" / f"ib_report_{broker}_{stamp}.xlsx"
    write_report_xlsx(result, report_path)
    return result, report_path


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python tracker.py file.xlsx")
        sys.exit(1)
    src = Path(sys.argv[1])
    broker = detect_broker(src.name)
    if not broker:
        print("Put PU Prime or Vantage in the file name.")
        sys.exit(1)
    out, report = process_file(src, broker=broker)
    print(format_telegram(out, src.name))
    print(f"\nReport: {report}")
